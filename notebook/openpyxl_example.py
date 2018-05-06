import openpyxl
import pprint

wb = openpyxl.load_workbook('data/src/sample.xlsx')

print(type(wb))
# <class 'openpyxl.workbook.workbook.Workbook'>

print(wb.sheetnames)
# ['sheet1', 'sheet2']

sheet = wb['sheet1']

print(type(sheet))
# <class 'openpyxl.worksheet.worksheet.Worksheet'>

cell = sheet['A2']

print(type(cell))
# <class 'openpyxl.cell.cell.Cell'>

print(cell.value)
# one

cell = sheet.cell(row=2, column=1)

print(type(cell))
# <class 'openpyxl.cell.cell.Cell'>

print(cell.value)
# one

pprint.pprint(sheet['A2:C4'])
# ((<Cell 'sheet1'.A2>, <Cell 'sheet1'.B2>, <Cell 'sheet1'.C2>),
#  (<Cell 'sheet1'.A3>, <Cell 'sheet1'.B3>, <Cell 'sheet1'.C3>),
#  (<Cell 'sheet1'.A4>, <Cell 'sheet1'.B4>, <Cell 'sheet1'.C4>))

g = sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3)

print(type(g))
# <class 'generator'>

pprint.pprint(list(g))
# [(<Cell 'sheet1'.A2>, <Cell 'sheet1'.B2>, <Cell 'sheet1'.C2>),
#  (<Cell 'sheet1'.A3>, <Cell 'sheet1'.B3>, <Cell 'sheet1'.C3>),
#  (<Cell 'sheet1'.A4>, <Cell 'sheet1'.B4>, <Cell 'sheet1'.C4>)]

def get_value_list(t_2d):
    return([[cell.value for cell in row] for row in t_2d])

l_2d = get_value_list(sheet['A2:C4'])

pprint.pprint(l_2d, width=40)
# [['one', 11.0, 12.0],
#  ['two', 21.0, 22.0],
#  ['three', 31.0, 32.0]]

def get_list_2d(sheet, start_row, end_row, start_col, end_col):
    return get_value_list(sheet.iter_rows(min_row=start_row,
                                          max_row=end_row,
                                          min_col=start_col,
                                          max_col=end_col))

l_2d = get_list_2d(sheet, 2, 4, 1, 3)

pprint.pprint(l_2d, width=40)
# [['one', 11.0, 12.0],
#  ['two', 21.0, 22.0],
#  ['three', 31.0, 32.0]]

g_all = sheet.values

print(type(g_all))
# <class 'generator'>

pprint.pprint(list(g_all), width=40)
# [(None, 'A', 'B', 'C'),
#  ('one', 11.0, 12.0, 13.0),
#  ('two', 21.0, 22.0, 23.0),
#  ('three', 31.0, 32.0, 33.0)]

sheet['C1'] = 'XXX'
sheet['E1'] = 'new'

pprint.pprint(list(sheet.values), width=40)
# [(None, 'A', 'XXX', 'C', 'new'),
#  ('one', 11.0, 12.0, 13.0, None),
#  ('two', 21.0, 22.0, 23.0, None),
#  ('three', 31.0, 32.0, 33.0, None)]

sheet.cell(row=2, column=5, value=14)

pprint.pprint(list(sheet.values), width=40)
# [(None, 'A', 'XXX', 'C', 'new'),
#  ('one', 11.0, 12.0, 13.0, 14),
#  ('two', 21.0, 22.0, 23.0, None),
#  ('three', 31.0, 32.0, 33.0, None)]

def write_list_2d(sheet, l_2d, start_row, start_col):
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])

l_2d = [['four', 41, 42, 43], ['five', 51, 52, 53]]

write_list_2d(sheet, l_2d, 5, 1)

pprint.pprint(list(sheet.values), width=40)
# [(None, 'A', 'XXX', 'C', 'new'),
#  ('one', 11.0, 12.0, 13.0, 14),
#  ('two', 21.0, 22.0, 23.0, None),
#  ('three', 31.0, 32.0, 33.0, None),
#  ('four', 41, 42, 43, None),
#  ('five', 51, 52, 53, None)]

sheet_new = wb.create_sheet('sheet_new')

print(wb.worksheets)
# [<Worksheet "sheet1">, <Worksheet "sheet2">, <Worksheet "sheet_new">]

sheet_new['A1'] = 'new sheet!'

print(list(sheet_new.values))
# [('new sheet!',)]

sheet_copy = wb.copy_worksheet(wb['sheet1'])

print(wb.worksheets)
# [<Worksheet "sheet1">, <Worksheet "sheet2">, <Worksheet "sheet_new">, <Worksheet "sheet1 Copy">]

pprint.pprint(list(sheet_copy.values))
# [(None, 'A', 'XXX', 'C', 'new'),
#  ('one', 11.0, 12.0, 13.0, 14),
#  ('two', 21.0, 22.0, 23.0, None),
#  ('three', 31.0, 32.0, 33.0, None),
#  ('four', 41, 42, 43, None),
#  ('five', 51, 52, 53, None)]

wb.remove_sheet(wb['sheet1 Copy'])

print(wb.worksheets)
# [<Worksheet "sheet1">, <Worksheet "sheet2">, <Worksheet "sheet_new">]

wb.save('data/dst/openpyxl_sample.xlsx')
